import time
import re
import os
import pandas as pd
from datetime import datetime
from ppadb.client import Client as AdbClient
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side

# ADB 연결
client = AdbClient(host="127.0.0.1", port=5037)

# 파일 경로
log_path = f'bluetooth_connect_report_{datetime.now().strftime("%Y%m%d_%H%M")}.txt'
summary_path = f'bluetooth_summary_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

connected_seconds = 0
disconnected_seconds = 0
interval = 10

# 실행 시작 시각
start_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
device_model = "Unknown"
connected_names = []

try:
    with open(log_path, mode='w', encoding='utf-8') as file:
        while True:
            devices = client.devices()
            if not devices:
                device_status = "연결되지 않음"
                bluetooth_status = "N/A"
                connected = False
            else:
                device_status = "연결됨"
                device = devices[0]
                device_model = device.shell("getprop ro.product.model").strip()
                output = device.shell("dumpsys bluetooth_manager")

                connected_devices = re.findall(r"\(Connected\)\s+([0-9A-F:]{17})", output)
                name_map = {}
                db_entries = re.findall(r"([0-9A-F:]{17})\s+\|\s+(.+?)\s+\|", output)
                for mac, name in db_entries:
                    name_map[mac] = name.strip()

                connected_names = []
                for mac in connected_devices:
                    name = name_map.get(mac, "기기명 불명")
                    connected_names.append(name)

                if connected_names:
                    bluetooth_status = f"이어폰 연결됨 ({', '.join(connected_names)})"
                    connected = True
                else:
                    bluetooth_status = "이어폰 미연결"
                    connected = False

            timestamp = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
            file.write(f"{timestamp} - 디바이스 상태: {device_status} - 블루투스 상태: {bluetooth_status}\n")
            file.flush()

            if connected:
                connected_seconds += interval
            else:
                disconnected_seconds += interval

            time.sleep(interval)

except KeyboardInterrupt:
    end_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    total_time = connected_seconds + disconnected_seconds
    connected_ratio = round((connected_seconds / total_time) * 100, 1) if total_time else 0
    disconnected_ratio = 100 - connected_ratio
    result = "PASS" if connected_ratio >= 80 else "FAIL"

    # 요약 데이터프레임
    summary = pd.DataFrame([
        {"상태": "실행 시작", "누적 시간 (초)": start_time, "비율 (%)": ""},
        {"상태": "실행 종료", "누적 시간 (초)": end_time, "비율 (%)": ""},
        {"상태": "연결됨", "누적 시간 (초)": connected_seconds, "비율 (%)": connected_ratio},
        {"상태": "미연결", "누적 시간 (초)": disconnected_seconds, "비율 (%)": disconnected_ratio},
        {"상태": "전체 실행 시간", "누적 시간 (초)": total_time, "비율 (%)": 100.0},
        {"상태": "판정 결과", "누적 시간 (초)": "", "비율 (%)": result}
    ])
    summary.to_excel(summary_path, index=False)

    # 엑셀 불러오기
    wb = load_workbook(summary_path)
    ws = wb.active

    # 상단 정보 삽입
    ws.insert_rows(1, amount=3)
    ws["A1"] = f"디바이스 이름: {device_model}"
    ws["A2"] = f"연결된 이어폰: {', '.join(connected_names) if connected_names else '없음'}"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"].font = Font(size=14, bold=True)

    # 열 너비 + 정렬 설정
    for col_idx, col_name in enumerate(["상태", "누적 시간 (초)", "비율 (%)"], start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20
        align = Alignment(horizontal="left") if col_name == "상태" else Alignment(horizontal="right")
        for cell in ws[col_letter][4:]:  # 4행부터 데이터 시작
            cell.alignment = align

    # 판정 결과 셀 Bold + 색상
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == "판정 결과":
            result_cell = row[2]
            result_cell.font = Font(bold=True, color="0000FF" if result_cell.value == "PASS" else "FF0000")
            break

    # 테두리 설정
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    min_row = 4
    max_row = ws.max_row
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border

    wb.save(summary_path)
    print(f"\n 리포트 저장 완료: {summary_path}")
    os.system(f"start excel {summary_path}")
